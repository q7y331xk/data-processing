from openpyxl import Workbook
from config import EXCEL_SAVE_PATH

def write_excel(sellings):
    col_length = len(sellings)
    row_length = len(sellings[0])
    wb = Workbook()
    ws = wb.active
    ws.title = "sellings"
    ws.cell(row = 1, column = 1, value = "id")
    ws.cell(row = 1, column = 2, value = "title")
    ws.cell(row = 1, column = 3, value = "cost")
    ws.cell(row = 1, column = 4, value = "nickname")
    ws.cell(row = 1, column = 5, value = "status")
    ws.cell(row = 1, column = 6, value = "use_cnt")
    ws.cell(row = 1, column = 7, value = "condition")
    ws.cell(row = 1, column = 8, value = "pay_methods")
    ws.cell(row = 1, column = 9, value = "delivery")
    ws.cell(row = 1, column = 10, value = "location")
    ws.cell(row = 1, column = 11, value = "main")
    ws.cell(row = 1, column = 12, value = "views")
    ws.cell(row = 1, column = 13, value = "date")
    ws.cell(row = 1, column = 14, value = "likes")
    ws.cell(row = 1, column = 15, value = "comments_cnt")
    ws.cell(row = 1, column = 16, value = "comments")
    ws.cell(row = 1, column = 17, value = "div")
    ws.cell(row = 1, column = 18, value = "gu")
    ws.cell(row = 1, column = 19, value = "color")
    i = 0
    while (i < col_length):
        row = sellings[i]
        j = 0
        while (j < row_length):
            ws.cell(row = i + 1 + 1, column = j + 1, value = row[j])
            j = j + 1
        i = i + 1
    wb.save(EXCEL_SAVE_PATH)
    print(f'write excel in "{EXCEL_SAVE_PATH}"')